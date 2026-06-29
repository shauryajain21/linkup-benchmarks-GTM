import csv, re, json, os, sys, random
from collections import OrderedDict, defaultdict

HERE  = os.path.dirname(os.path.abspath(__file__))
JC    = os.path.join(HERE, "data", "jc_outputs_57.csv")
GTF   = os.path.join(HERE, "data", "job_change_ground_truth_57.csv")
RAW   = os.path.join(HERE, "results", "jc_judge_raw.json")
XLSX  = os.path.join(HERE, "results", "freshness_judge.xlsx")
PROVS = ["linkup", "exa", "perplexity", "parallel"]
CLASSES = ["FRESH", "STALE", "WRONG_PERSON", "NOT_FOUND"]
JUDGE_MODEL = "claude-opus-4-8"

LABELS = ["engine_A", "engine_B", "engine_C", "engine_D"]

def blind_map(key):

    order = list(PROVS)
    random.Random(f"blind::{key}").shuffle(order)
    lab2eng = {LABELS[i]: e for i, e in enumerate(order)}
    return lab2eng, {e: l for l, e in lab2eng.items()}

def name_q(q): m = re.search(r"Research ([^(]+)\(", q); return m.group(1).strip() if m else ""
def jload(s):
    s = (s or "").strip()
    try: return json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, re.S)
        try: return json.loads(m.group(0)) if m else {}
        except Exception: return {}

def load():
    gt = {r["name"].strip().lower(): r for r in csv.DictReader(open(GTF))}
    people = OrderedDict()
    for r in csv.DictReader(open(JC)):
        nm = name_q(r["query"]).lower()
        people.setdefault(nm, {"name": name_q(r["query"]), "resp": {}})
        people[nm]["resp"][r["provider"]] = jload(r["response"])
    return gt, people

SCHEMA = {"type": "object", "properties": {p: {"type": "object", "properties": {
    "classification": {"type": "string", "enum": CLASSES},
    "detected_change": {"type": "boolean"},
    "note": {"type": "string"}},
    "required": ["classification", "detected_change", "note"], "additionalProperties": False} for p in LABELS}}
SCHEMA["required"] = LABELS; SCHEMA["additionalProperties"] = False

SYS = ("You evaluate JOB-CHANGE FRESHNESS. Each target person RECENTLY changed jobs. You are given the "
 "verified ground truth: their NEW current company/title (and when it started) and their PREVIOUS company. "
 "Company names may appear in different languages, transliterations, or abbreviations across sources — "
 "treat 'JP Morgan'='J.P. Morgan', and an Arabic company name = its English/transliterated equivalent as a MATCH. "
 "For each engine, look at the current_company / current_title / change_summary it returned and classify:\n"
 "- FRESH: it reported the NEW current company (caught the recent change). The single most important outcome.\n"
 "- STALE: it found the right person but reported the PREVIOUS/older company (missed the change).\n"
 "- WRONG_PERSON: it returned a different individual (namesake) or a company matching neither new nor previous.\n"
 "- NOT_FOUND: it abstained / found=false / empty current company.\n"
 "Also set detected_change = did the engine correctly assert the person changed companies (job_changed true "
 "AND consistent with moving to the new role). Be strict; base it on the ground truth, not the engine's confidence.\n"
 "- note: one short phrase.")

def run_judge(people, gt):
    from anthropic import Anthropic
    import asyncio
    client = Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    results = {}
    if os.path.exists(RAW):
        for nm, rec in json.load(open(RAW)).items():
            j = rec.get("judge", {})
            if isinstance(j, dict) and "_err" not in j: results[nm] = j
    todo = [(nm, p) for nm, p in people.items() if nm not in results]
    print(f"judging: {len(results)} cached, {len(todo)} to run")

    def judge(name, g, resp):
        lab2eng, eng2lab = blind_map(name)
        payload = {"person": name,
            "ground_truth": {"new_current_company": g.get("current_company"), "new_current_title": g.get("current_title"),
                             "new_role_start": g.get("current_role_start"),
                             "previous_company": g.get("previous_company") or g.get("snapshot_company"),
                             "previous_title": g.get("previous_title") or g.get("snapshot_title")},
            "engine_outputs": {lab: {k: resp.get(lab2eng[lab], {}).get(k) for k in
                ["found", "current_company", "current_title", "current_start_date",
                 "previous_company", "job_changed", "change_summary"]} for lab in LABELS}}
        msg = client.messages.create(model=JUDGE_MODEL, max_tokens=1500, system=SYS,
            messages=[{"role": "user", "content": json.dumps(payload, ensure_ascii=False)[:60000] +
                       "\n\nReturn ONLY JSON matching this schema:\n" + json.dumps(SCHEMA)}])
        verdict = jload(msg.content[0].text)

        if isinstance(verdict, dict):
            return {lab2eng[lab]: verdict[lab] for lab in LABELS if lab in verdict}
        return verdict

    async def main():
        loop = asyncio.get_event_loop(); sem = asyncio.Semaphore(8); done = [len(results)]
        async def one(nm, p):
            async with sem:
                try:
                    v = await loop.run_in_executor(None, judge, p["name"], gt.get(nm, {}), p["resp"])
                    results[nm] = v; done[0] += 1; print(f"[{done[0]}/{len(people)}] {p['name']}")
                except Exception as e:
                    results[nm] = {"_err": str(e)}; print("ERR", p["name"], repr(e))
        await asyncio.gather(*[one(nm, p) for nm, p in todo])
    asyncio.run(main())
    os.makedirs(os.path.dirname(RAW), exist_ok=True)
    json.dump({nm: {"name": people[nm]["name"], "judge": results[nm]} for nm in people},
              open(RAW, "w"), indent=1, ensure_ascii=False)

def score_and_report():
    gt, people = load()
    raw = json.load(open(RAW))
    agg = {p: defaultdict(int) for p in PROVS}; det = {p: 0 for p in PROVS}; n = 0; rows = []
    for nm, rec in raw.items():
        j = rec.get("judge", {})
        if not isinstance(j, dict) or "_err" in j: continue
        n += 1; g = gt.get(nm, {}); resp = people.get(nm, {}).get("resp", {})
        row = {"name": rec["name"], "new": g.get("current_company", ""), "prev": g.get("previous_company") or g.get("snapshot_company", "")}
        for p in PROVS:
            v = j.get(p)
            if not isinstance(v, dict): row[p] = ("", "?"); continue
            c = v.get("classification", "?"); agg[p][c] += 1; det[p] += int(v.get("detected_change") is True)
            row[p] = (resp.get(p, {}).get("current_company", ""), c)
        rows.append(row)

    print(f"\n=== JOB-CHANGE FRESHNESS — Opus judge, ground-truth (n={n}) ===\n")
    print(f"{'engine':11}{'FRESH':>7}{'STALE':>7}{'WRONG':>7}{'NOTFND':>7}{'fresh%':>8}{'chg✓':>6}")
    for p in sorted(PROVS, key=lambda x: -agg[x]['FRESH']):
        a = agg[p]
        print(f"{p:11}{a['FRESH']:>7}{a['STALE']:>7}{a['WRONG_PERSON']:>7}{a['NOT_FOUND']:>7}{100*a['FRESH']/n:>7.0f}%{det[p]:>6}")
    print("\nFRESH = reported the NEW company (caught the change) | STALE = old company | chg✓ = correctly flagged a change")
    write_excel(agg, det, n, rows)
    print(f"SAVED {XLSX}")

def write_excel(agg, det, n, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); HDR = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="2F5496")
    FILL = {"FRESH": PatternFill("solid", fgColor="C6EFCE"), "STALE": PatternFill("solid", fgColor="FFEB9C"),
            "WRONG_PERSON": PatternFill("solid", fgColor="FFC7CE"), "NOT_FOUND": PatternFill("solid", fgColor="E0E0E0")}
    ws = wb.active; ws.title = "Summary"
    ws.append(["Job-change freshness — did the engine surface the NEW company? (n=%d)" % n]); ws["A1"].font = Font(bold=True, size=12)
    ws.append(["engine", "FRESH", "STALE", "WRONG_PERSON", "NOT_FOUND", "fresh %", "change flagged ✓"])
    for c in ws[2]: c.font = HDR; c.fill = HF
    for p in sorted(PROVS, key=lambda x: -agg[x]['FRESH']):
        a = agg[p]
        ws.append([p, a['FRESH'], a['STALE'], a['WRONG_PERSON'], a['NOT_FOUND'], round(100*a['FRESH']/n), det[p]])
        ws.cell(ws.max_row, 2).fill = FILL["FRESH"]
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFG": ws.column_dimensions[col].width = 15

    wp = wb.create_sheet("Per-person")
    hdr = ["Person", "NEW company (truth)", "Previous (truth)"]
    for p in PROVS: hdr += [p + " returned", p]
    wp.append(hdr)
    for c in wp[1]: c.font = HDR; c.fill = HF
    for row in rows:
        line = [row["name"], row["new"], row["prev"]]
        for p in PROVS: line += [row[p][0], row[p][1]]
        wp.append(line); r = wp.max_row
        for i, p in enumerate(PROVS):
            cls = row[p][1]; cell = wp.cell(r, 5 + 2*i)
            if cls in FILL: cell.fill = FILL[cls]
            cell.alignment = Alignment(horizontal="center")
    wp.freeze_panes = "A2"
    wp.column_dimensions["A"].width = 22; wp.column_dimensions["B"].width = 28; wp.column_dimensions["C"].width = 24
    for i in range(len(PROVS)):
        from openpyxl.utils import get_column_letter
        wp.column_dimensions[get_column_letter(4 + 2*i)].width = 26
        wp.column_dimensions[get_column_letter(5 + 2*i)].width = 13
    os.makedirs(os.path.dirname(XLSX), exist_ok=True); wb.save(XLSX)

if __name__ == "__main__":
    if "--rejudge" in sys.argv and os.path.exists(RAW): os.remove(RAW)
    if not os.path.exists(RAW):
        gt, people = load(); run_judge(people, gt)
    score_and_report()
