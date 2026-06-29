import csv, re, json, os
from collections import OrderedDict, defaultdict
from statistics import mean, median
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUTS = os.path.join(HERE, "data", "api_outputs_500.csv")
TRUTH   = os.path.join(HERE, "data", "ground_truth_500.csv")
RESULT  = os.path.join(HERE, "results", "benchmark_500.xlsx")
PROVS   = ["linkup", "exa", "perplexity", "parallel"]

PLACE = re.compile(r"^\s*$|not found|not provided|not available|unknown|^n/?a$|"
                   r"no information|not specified|not listed|do not contain", re.I)

def filled(v):
    if v is None: return False
    if isinstance(v, (list, dict)): return len(v) > 0
    s = str(v).strip()
    return bool(s) and not PLACE.search(s)

def jload(s):

    s = (s or "").strip()
    if "```" in s:
        for p in s.split("```"):
            p = p.strip().lstrip("json").strip()
            if p.startswith("{"): s = p; break
    try:
        return json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, re.S)
        try:
            return json.loads(m.group(0)) if m else {}
        except Exception:
            return {}

def norm_url(u):
    u = (u or "").strip().lower().split("?")[0].rstrip("/")
    return re.sub(r"^https?://(www\.)?", "", u)

def url_of(q):
    m = re.search(r"https?://[^\s)]+", q); return m.group(0) if m else ""

def name_of(q):
    m = re.search(r"\(([^()]+)\)", q); return m.group(1).strip() if m else ""

SUFFIX = re.compile(r"\b(inc|llc|ltd|plc|corp|corporation|company|co|gmbh|sa|pvt|"
                    r"private|limited|group|technologies|technology|solutions|the|at|ex)\b", re.I)
def cnorm(s):
    s = re.sub(r"[^a-z0-9 ]", " ", (s or "").lower()); s = SUFFIX.sub(" ", s)
    return set(w for w in s.split() if len(w) > 1)

def match(a, b):

    if not filled(a) or not filled(b): return None
    A, B = cnorm(a), cnorm(b)
    if not A or not B: return None
    if A & B and (A <= B or B <= A): return True
    return len(A & B) / len(A | B) >= 0.34

FIELDS = ["full_name", "location", "headline", "current_company", "current_title",
          "experience", "exp_dates", "education", "recent_posts", "post_engagement"]

def completeness(d):
    if not isinstance(d, dict): return 0, {f: 0 for f in FIELDS}
    p = {}
    for k in ["full_name", "location", "headline", "current_company", "current_title"]:
        p[k] = int(filled(d.get(k)))
    exp = d.get("experience") or []
    p["experience"] = int(len(exp) > 0)
    p["exp_dates"]  = int(any(filled(e.get("start")) for e in exp if isinstance(e, dict)))
    p["education"]  = int(len(d.get("education") or []) > 0)
    posts = d.get("recent_posts") or d.get("posts") or []
    p["recent_posts"]    = int(len(posts) > 0)
    p["post_engagement"] = int(any(("like" in str(k).lower() or "comment" in str(k).lower()) and filled(v)
                                   for pp in posts if isinstance(pp, dict) for k, v in pp.items()))
    return round(100 * sum(p.values()) / len(p)), p

def load():
    gt = {norm_url(r["linkedin_url"]): r for r in csv.DictReader(open(TRUTH))}
    people = OrderedDict()
    for r in csv.DictReader(open(OUTPUTS)):
        u = norm_url(url_of(r["query"]))
        people.setdefault(u, {"name": name_of(r["query"]), "resp": {}})
        people[u]["resp"][r["provider"]] = jload(r["response"])
    return gt, people

def run():
    gt, people = load()
    matched = sum(1 for u in people if u in gt)
    print(f"people: {len(people)} | matched to ground-truth DB: {matched}\n")

    comp = defaultdict(list)
    fagg = {p: defaultdict(int) for p in PROVS}
    cnt  = defaultdict(int)
    corr = {p: defaultdict(int) for p in PROVS}
    detail = []

    for u, pp in people.items():
        g = gt.get(u)
        row = {"name": pp["name"], "url": u,
               "gt_company": g["current_company"] if g else "",
               "gt_title":   g["current_title"]   if g else ""}
        for p in PROVS:
            d = pp["resp"].get(p, {})

            sc, pts = completeness(d)
            comp[p].append(sc); cnt[p] += 1
            for f in FIELDS: fagg[p][f] += pts[f]
            row[p + "_comp"] = sc

            co = d.get("current_company", ""); ti = d.get("current_title", ""); nm = d.get("full_name", "")
            if g:
                cm = match(co, g["current_company"]); tm = match(ti, g["current_title"]); nmm = match(nm, g["name"])
                corr[p]["n"] += 1
                if not filled(co):  corr[p]["empty"] += 1; flag = "—"
                elif cm:            corr[p]["right"] += 1; flag = "✓"
                else:               corr[p]["wrong"] += 1; flag = "✗"
                if tm:  corr[p]["ti"] += 1
                if nmm: corr[p]["nm"] += 1
                row[p + "_company"] = co; row[p + "_match"] = flag
            else:
                row[p + "_company"] = co; row[p + "_match"] = "?"
        detail.append(row)

    print("=== COMPLETENESS (0-100) ===")
    print(f"{'provider':11}{'mean':>7}{'median':>8}{'empty0':>8}")
    for p in sorted(PROVS, key=lambda x: -mean(comp[x])):
        s = comp[p]
        print(f"{p:11}{mean(s):7.1f}{median(s):8.0f}{sum(1 for x in s if x == 0):8}")

    print("\n=== CORRECTNESS vs ground-truth DB ===")
    print(f"{'provider':11}{'right%':>8}{'wrong':>7}{'empty':>7}{'title%':>8}{'name%':>7}")
    for p in sorted(PROVS, key=lambda x: -corr[x]['right']):
        a = corr[p]; n = a['n']
        print(f"{p:11}{100*a['right']/n:>7.0f}%{a['wrong']:>7}{a['empty']:>7}{100*a['ti']/n:>7.0f}%{100*a['nm']/n:>6.0f}%")

    write_excel(comp, corr, fagg, cnt, detail)
    print(f"\nSAVED {RESULT}")

def write_excel(comp, corr, fagg, cnt, detail):
    wb = Workbook()
    HDR = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="2F5496")
    GREEN = PatternFill("solid", fgColor="C6EFCE"); YEL = PatternFill("solid", fgColor="FFEB9C")
    ORA = PatternFill("solid", fgColor="FFD0B0"); RED = PatternFill("solid", fgColor="FFC7CE")
    GREY = PatternFill("solid", fgColor="E0E0E0")
    def hdrify(ws):
        for c in ws[1]:
            c.font = HDR; c.fill = HF; c.alignment = Alignment(horizontal="center", vertical="center")
    def cscale(c, v):
        c.fill = GREEN if v >= 85 else YEL if v >= 60 else ORA if v > 0 else RED

    ws = wb.active; ws.title = "Summary"
    ws.append(["COMPLETENESS"]); ws["A1"].font = Font(bold=True, size=13)
    ws.append(["provider", "mean", "median", "empty(0)", "n"])
    for p in sorted(PROVS, key=lambda x: -mean(comp[x])):
        s = comp[p]; ws.append([p, round(mean(s), 1), median(s), sum(1 for x in s if x == 0), len(s)])
    ws.append([])
    ws.append(["CORRECTNESS vs ground-truth DB"]); ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
    ws.append(["provider", "right-person %", "wrong (namesake)", "empty", "title %", "name %", "n"])
    for p in sorted(PROVS, key=lambda x: -corr[x]['right']):
        a = corr[p]; n = a['n']
        ws.append([p, round(100*a['right']/n), a['wrong'], a['empty'], round(100*a['ti']/n), round(100*a['nm']/n), n])
    ws.append([])
    ws.append(["per-field fill rate (%)"]); ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append(["field"] + PROVS)
    for f in FIELDS:
        ws.append([f] + [round(100*fagg[p][f]/cnt[p]) for p in PROVS])
    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 18 if col == "A" else 16

    wc = wb.create_sheet("Completeness")
    wc.append(["Person"] + [p.capitalize() for p in PROVS] + ["Best"])
    for row in detail:
        sc = {p: row[p + "_comp"] for p in PROVS}; best = max(sc, key=sc.get)
        wc.append([row["name"]] + [sc[p] for p in PROVS] + [best.capitalize()])
        for i, p in enumerate(PROVS): cscale(wc.cell(wc.max_row, 2 + i), sc[p])
    hdrify(wc); wc.freeze_panes = "A2"; wc.column_dimensions["A"].width = 26
    for col in range(2, 7): wc.column_dimensions[get_column_letter(col)].width = 12

    wm = wb.create_sheet("Correctness detail")
    hdr = ["Person", "TRUTH company", "TRUTH title"]
    for p in PROVS: hdr += [p, "OK"]
    wm.append(hdr)
    for row in detail:
        line = [row["name"], row["gt_company"], row["gt_title"]]
        for p in PROVS: line += [row[p + "_company"], row[p + "_match"]]
        wm.append(line); r = wm.max_row
        for i, p in enumerate(PROVS):
            c = wm.cell(r, 5 + 2*i); m = row[p + "_match"]
            c.fill = GREEN if m == "✓" else RED if m == "✗" else GREY
            c.alignment = Alignment(horizontal="center")
    hdrify(wm); wm.freeze_panes = "A2"
    wm.column_dimensions["A"].width = 24; wm.column_dimensions["B"].width = 26; wm.column_dimensions["C"].width = 22
    for i in range(len(PROVS)):
        wm.column_dimensions[get_column_letter(4 + 2*i)].width = 28
        wm.column_dimensions[get_column_letter(5 + 2*i)].width = 4

    os.makedirs(os.path.dirname(RESULT), exist_ok=True)
    wb.save(RESULT)

if __name__ == "__main__":
    run()
