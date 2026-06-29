import csv, re, json, os, sys, random
from collections import OrderedDict, defaultdict
from statistics import mean

HERE     = os.path.dirname(os.path.abspath(__file__))
BRIEFS   = os.path.join(HERE, "data", "brief_outputs_100.csv")
TRUTH    = os.path.join(HERE, "data", "ground_truth_500.csv")
RAW      = os.path.join(HERE, "results", "brief_judge_raw.json")
XLSX     = os.path.join(HERE, "results", "meeting_brief_judge.xlsx")
PROVS    = ["linkup", "exa", "perplexity", "parallel"]
JUDGE_MODEL = "claude-opus-4-8"

LABELS = ["engine_A", "engine_B", "engine_C", "engine_D"]

def blind_map(key):

    order = list(PROVS)
    random.Random(f"blind::{key}").shuffle(order)
    lab2eng = {LABELS[i]: e for i, e in enumerate(order)}
    return lab2eng, {e: l for l, e in lab2eng.items()}

def norm_url(u):
    u = (u or "").strip().lower().split("?")[0].rstrip("/")
    return re.sub(r"^https?://(www\.)?", "", u)
def url_of(q):  m = re.search(r"https?://[^\s)]+", q); return m.group(0) if m else ""
def name_of(q): m = re.search(r"\(([^()]+)\)", q);     return m.group(1).strip() if m else ""
def jload(s):
    s = (s or "").strip()
    try: return json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, re.S)
        try: return json.loads(m.group(0)) if m else {}
        except Exception: return {}
def n0(x):
    try: return float(x)
    except Exception: return 0.0

def overall(v):

    o = 0.35 * n0(v.get("freshness")) + 0.30 * n0(v.get("specificity")) + 0.35 * n0(v.get("actionability"))
    if v.get("right_person") is False:
        o = min(15.0, o)
    return round(o, 1)

def load():
    gt = {norm_url(r["linkedin_url"]): r for r in csv.DictReader(open(TRUTH))}
    people = OrderedDict()
    for r in csv.DictReader(open(BRIEFS)):
        u = norm_url(url_of(r["query"]))
        people.setdefault(u, {"name": name_of(r["query"]), "briefs": {}})
        people[u]["briefs"][r["provider"]] = {
            "about": r.get("about_target"), "n_notes": r.get("n_notes"), "n_q": r.get("n_questions"),
            "notes": (r.get("notes") or "")[:4000], "questions": (r.get("questions") or "")[:2000]}
    return gt, people

SCHEMA = {"type": "object", "properties": {p: {"type": "object", "properties": {
    "right_person":  {"type": "boolean"},
    "freshness":     {"type": "integer"},
    "specificity":   {"type": "integer"},
    "actionability": {"type": "integer"},
    "overall":       {"type": "integer"},
    "note":          {"type": "string"}},
    "required": ["right_person", "freshness", "specificity", "actionability", "overall", "note"],
    "additionalProperties": False} for p in LABELS}}
SCHEMA["required"] = LABELS; SCHEMA["additionalProperties"] = False

SYS = (
 "ROLE. You are a salesperson about to walk into a meeting with the target person in 10 minutes. "
 "You have several briefs in front of you (one per engine), each with notes + suggested conversation "
 "questions. You also know the target's VERIFIED current company/title (ground truth). Judge each brief "
 "by one test: which one actually prepares ME better to open a great conversation right now?\n\n"
 "WHAT A GREAT BRIEF LOOKS LIKE (in priority order):\n"
 "1. RECENT & TIME-RELEVANT beats old history. A post from the last few weeks, a just-announced role "
 "change, a current initiative or event I can mention TODAY is worth far more than a 2019 job date or a "
 "static bio. Reward freshness; a brief full of only stale career history is weak even if detailed.\n"
 "2. SPECIFIC & TRUE beats vague. Concrete names, numbers, dates, deals, products I can name out loud. "
 "But specificity only counts if it's GROUNDED — anything that contradicts ground truth, invents future "
 "dates, or looks fabricated SUBTRACTS (a confident wrong fact is worse than no fact, it'll embarrass me).\n"
 "3. SHARP QUESTIONS beat boilerplate. Questions tailored to THIS person's actual situation ('I saw you "
 "just moved X to Y — how's that going?') are gold. Generic filler ('What are your goals this year?') "
 "that could be asked of anyone adds almost nothing.\n\n"
 "SCORE EACH ENGINE (be strict and discriminating — do NOT cluster everything at 80; spread the scores):\n"
 "- right_person (bool): is this brief truly about the target (matches ground-truth company/trajectory, "
 "OR a plausibly more-recent role for the same individual), not a namesake? An empty brief = false. "
 "A fresher current company than ground truth is NOT a wrong person.\n"
 "- freshness 0-100: how much recent, time-relevant intel I can use today. Mostly-stale = low. Empty = 0.\n"
 "- specificity 0-100: density of concrete, GROUNDED detail. Vague filler = low; fabricated/contradictory "
 "claims pull it DOWN. Empty = 0.\n"
 "- actionability 0-100: are the questions sharp and tailored vs generic? Empty/boilerplate = low.\n"
 "- overall 0-100: (we recompute this ourselves; still provide your best estimate.)\n"
 "- note: one short phrase on why.")

def run_judge(people, gt):
    from anthropic import Anthropic
    import asyncio
    client = Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    results = {}
    if os.path.exists(RAW):
        for u, rec in json.load(open(RAW)).items():
            j = rec.get("judge", {})
            if isinstance(j, dict) and "_err" not in j:
                results[u] = j
    todo = [(u, p) for u, p in people.items() if u not in results]
    print(f"judging: {len(results)} cached, {len(todo)} to run")

    def judge(name, u, g, briefs):
        lab2eng, eng2lab = blind_map(u)
        payload = {"target": name, "linkedin_url": u,
                   "ground_truth": {"current_company": g.get("current_company"), "current_title": g.get("current_title"),
                                    "headline": g.get("headline"), "location": g.get("location")},
                   "briefs": {lab: briefs.get(lab2eng[lab], {}) for lab in LABELS}}
        msg = client.messages.create(model=JUDGE_MODEL, max_tokens=2000, system=SYS,
            messages=[{"role": "user", "content": json.dumps(payload)[:120000] +
                       "\n\nReturn ONLY JSON matching this schema:\n" + json.dumps(SCHEMA)}])
        verdict = jload(msg.content[0].text)

        if isinstance(verdict, dict):
            return {lab2eng[lab]: verdict[lab] for lab in LABELS if lab in verdict}
        return verdict

    async def main():
        loop = asyncio.get_event_loop(); sem = asyncio.Semaphore(8); done = [len(results)]
        async def one(u, p):
            async with sem:
                try:
                    v = await loop.run_in_executor(None, judge, p["name"], u, gt.get(u, {}), p["briefs"])
                    results[u] = v; done[0] += 1
                    print(f"[{done[0]}/{len(people)}] {p['name']}")
                except Exception as e:
                    results[u] = {"_err": str(e)}; print("ERR", p["name"], repr(e))
        await asyncio.gather(*[one(u, p) for u, p in todo])
    asyncio.run(main())
    os.makedirs(os.path.dirname(RAW), exist_ok=True)
    json.dump({u: {"name": people[u]["name"], "judge": results[u]} for u in people}, open(RAW, "w"), indent=1)
    return {u: results[u] for u in people}

def score_and_report():
    raw = json.load(open(RAW))
    agg = {p: {"overall": [], "freshness": [], "specificity": [], "actionability": []} for p in PROVS}
    right = {p: 0 for p in PROVS}; wins = {p: 0 for p in PROVS}; npp = 0; rows = []
    for u, rec in raw.items():
        j = rec.get("judge", {})
        if not isinstance(j, dict) or "_err" in j: continue
        npp += 1; best = None; bv = -1; row = {"name": rec["name"]}
        for p in PROVS:
            v = j.get(p)
            if not isinstance(v, dict): row[p] = None; continue
            o = overall(v); agg[p]["overall"].append(o)
            for d in ["freshness", "specificity", "actionability"]: agg[p][d].append(n0(v.get(d)))
            right[p] += int(v.get("right_person") is True)
            row[p] = {"overall": o, "f": n0(v.get("freshness")), "s": n0(v.get("specificity")),
                      "a": n0(v.get("actionability")), "note": v.get("note", "")}
            if o > bv: bv = o; best = p
        if best: wins[best] += 1
        rows.append(row)

    print(f"\n=== PRE-MEETING BRIEF JUDGE — salesperson rubric (n={npp}) ===\n")
    print(f"{'engine':11}{'OVERALL':>9}{'fresh':>7}{'specif':>7}{'action':>7}{'right%':>8}{'wins':>6}")
    for p in sorted(PROVS, key=lambda x: -mean(agg[x]['overall'])):
        a = agg[p]
        print(f"{p:11}{mean(a['overall']):>9.1f}{mean(a['freshness']):>7.1f}{mean(a['specificity']):>7.1f}"
              f"{mean(a['actionability']):>7.1f}{100*right[p]/npp:>7.0f}%{wins[p]:>6}")
    print("\noverall = freshness*.35 + specificity*.30 + actionability*.35, capped at 15 if wrong person")
    write_excel(agg, right, wins, npp, rows)
    print(f"SAVED {XLSX}")

def write_excel(agg, right, wins, npp, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); HDR = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="2F5496")
    GREEN = PatternFill("solid", fgColor="C6EFCE"); YEL = PatternFill("solid", fgColor="FFEB9C")
    ORA = PatternFill("solid", fgColor="FFD0B0"); RED = PatternFill("solid", fgColor="FFC7CE")
    def cs(c, v): c.fill = GREEN if v >= 65 else YEL if v >= 45 else ORA if v > 0 else RED

    ws = wb.active; ws.title = "Summary"
    ws.append(["Pre-meeting brief — salesperson rubric (freshness .35 / specificity .30 / actionability .35)"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(["engine", "OVERALL", "freshness", "specificity", "actionability", "right-person %", "wins (best/person)"])
    for c in ws[2]: c.font = HDR; c.fill = HF
    for p in sorted(PROVS, key=lambda x: -mean(agg[x]['overall'])):
        a = agg[p]
        ws.append([p, round(mean(a['overall']), 1), round(mean(a['freshness']), 1), round(mean(a['specificity']), 1),
                   round(mean(a['actionability']), 1), round(100*right[p]/npp), wins[p]])
        cs(ws.cell(ws.max_row, 2), mean(a['overall']))
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFG": ws.column_dimensions[col].width = 16

    wp = wb.create_sheet("Per-person")
    hdr = ["Person"]
    for p in PROVS: hdr += [p + " overall", p + " fresh"]
    hdr += ["Best"]; wp.append(hdr)
    for c in wp[1]: c.font = HDR; c.fill = HF
    for row in rows:
        ln = [row["name"]]; sc = {}
        for p in PROVS:
            v = row[p]
            if v: ln += [v["overall"], v["f"]]; sc[p] = v["overall"]
            else: ln += ["", ""]
        ln.append(max(sc, key=sc.get).capitalize() if sc else "-"); wp.append(ln)
        for i, p in enumerate(PROVS):
            if row[p]: cs(wp.cell(wp.max_row, 2 + 2*i), row[p]["overall"])
    wp.freeze_panes = "A2"; wp.column_dimensions["A"].width = 24
    os.makedirs(os.path.dirname(XLSX), exist_ok=True); wb.save(XLSX)

if __name__ == "__main__":
    rejudge = "--rejudge" in sys.argv
    if rejudge and os.path.exists(RAW): os.remove(RAW)
    if not os.path.exists(RAW):
        gt, people = load()
        run_judge(people, gt)
    score_and_report()
